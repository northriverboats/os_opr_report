#!/usr/bin/env python3
import re
import sys
import os
import xlwt
import click
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from hashlib import md5
from titlecase import titlecase
from emailer.emailer import Email
from mysql_tunnel.mysql_tunnel import TunnelSQL
from dotenv import load_dotenv
from pprint import pprint

fields = [
    'submitted',
    'dealership',
    'model',
    'hull_serial_number',
    'date_delivered',
    'agency',
    'first_name',
    'last_name',
    'phone_home',
    'email',
    'mailing_address',
    'mailing_city',
    'mailing_state',
    'mailing_zip',
]

states = {
    'Alaska': 'AK',
    'Alabama': 'AL',
    'Arkansas': 'AR',
    'American Samoa': 'AS',
    'Arizona': 'AZ',
    'California': 'CA',
    'Colorado': 'CO',
    'Connecticut': 'CT',
    'District of Columbia': 'DC',
    'Delaware': 'DE',
    'Florida': 'FL',
    'Georgia': 'GA',
    'Hawaii': 'HI',
    'Iowa': 'IA',
    'Idaho': 'ID',
    'Illinois': 'IL',
    'Indiana': 'IN',
    'Kansas': 'KS',
    'Kentucky': 'KY',
    'Louisiana': 'LA',
    'Massachusetts': 'MA',
    'Maryland': 'MD',
    'Maine': 'ME',
    'Michigan': 'MI',
    'Minnesota': 'MN',
    'Missouri': 'MO',
    'Mississippi': 'MS',
    'Montana': 'MT',
    'National': 'NA',
    'North Carolina': 'NC',
    'North Dakota': 'ND',
    'Nebraska': 'NE',
    'New Hampshire': 'NH',
    'New Jersey': 'NJ',
    'New Mexico': 'NM',
    'Nevada': 'NV',
    'New York': 'NY',
    'Ohio': 'OH',
    'Oklahoma': 'OK',
    'Oregon': 'OR',
    'Pennsylvania': 'PA',
    'Puerto Rico': 'PR',
    'Rhode Island': 'RI',
    'South Carolina': 'SC',
    'South Dakota': 'SD',
    'Tennessee': 'TN',
    'Texas': 'TX',
    'Utah': 'UT',
    'Virginia': 'VA',
    'Virgin Islands': 'VI',
    'Vermont': 'VT',
    'Washington': 'WA',
    'Wisconsin': 'WI',
    'West Virginia': 'WV',
    'Wyoming': 'WY'
}


"""
Levels
0 = no output
1 = minimal output
2 = verbose outupt
3 = very verbose outupt
"""
dbg = 0
def debug(level, text):
    if dbg > (level -1):
        print(text)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def resolve_flag(env_var, default):
    """convert enviromntal variable to True False
       return default value if no string"""
    if os.getenv(env_var):
        return [False, True][os.getenv(env_var) != ""]
    return default

def resolve_text(env_var, default):
    """convert enviromntal variable to text string
       return default value if no string"""
    if os.getenv(env_var):
        return os.getenv(env_var)
    return default

def resolve_int(env_var, default):
    return int(resolve_text(env_var, default))

def mail_results(subject, body, attachment=None):
    mFrom = os.getenv('MAIL_FROM')
    mTo = os.getenv('MAIL_TO')
    mCc = os.getenv('MAIL_CC')
    m = Email(os.getenv('MAIL_SERVER'))
    m.setFrom(mFrom)
    for email in mTo.split(','):
        m.addRecipient(email)
    for email in mCc.split(','):
        m.addCC(email)

    m.setSubject(subject)
    m.setTextBody("You should not see this text in a MIME aware reader")
    m.setHtmlBody(body)
    if (attachment):
        m.addAttachment(attachment)
    m.send()


def fetch_oprs(report_start, report_date):
    silent = dbg < 1
    with TunnelSQL(silent=silent, cursor='DictCursor') as db:
        # select all records from the OPR table
        sql = """
              SELECT  submitted, model, dealership, hull_serial_number, date_delivered,
                      agency, first_name,last_name, phone_home, email,
                      mailing_address, mailing_city, mailing_state, mailing_zip
                FROM  wp_nrb_opr
                WHERE  submitted BETWEEN %s and %s
            ORDER BY  submitted DESC
        """

        oprs = db.execute(sql, (report_start, report_date))

    return oprs



def write_sheet(oprs, xlsfile, report_date):
    wb = load_workbook(filename = xlsfile)
    ws = wb.active
    for row, opr in enumerate(oprs, start=2):
        opr['submitted'] = opr['submitted'].date()
        for column, field in enumerate(fields, start=1):
            _ = ws.cell(row=row, column=column, value=opr[field])
    title = report_date.strftime("%b %-d, %Y")
    filename = report_date.strftime("OPR Sales %Y-%m-%d.xlsx")
    longfilename = resource_path(filename)
    ws.title = title
    wb.save(filename = longfilename)
    return filename, longfilename

def opr_to_customer(opr):
    """create customre string from opr"""
    customer = ""
    if opr['agency']:
        customer += opr['agency'] + ', '
    customer += opr['first_name'] + ' ' + opr['last_name']
    return customer

def dump_opr(opr):
    print("{:10.10}  {:16.16}   {:20.20}  {:14.14}  {:10.10}  {:40.40}  "
          "{:14.14}  {:40.40}".format(
            opr['submitted'].strftime('%Y-%m-%d'),
            opr['dealership'],
            opr['model'],
            opr['hull_serial_number'],
            opr['date_delivered'].strftime('%Y-%m-%d'),
            opr_to_customer(opr),
            opr['phone_home'],
            opr['email'],
        ))

def dump_oprs(oprs):
    print ("----------  -----------------  --------------------  --------------  "
           "----------  ----------------------------------------  --------------  "
           "----------------------------------------")
    print("Submitted   Dealership         Model                 Serial "
          "Number   Delivered   Customer                                  "
          "Phone           Email")
    print ("----------  -----------------  --------------------  --------------  "
           "----------  ----------------------------------------  --------------  "
           "----------------------------------------")
    for opr in oprs:
        dump_opr(opr)


@click.command()
@click.option('--debug', '-d', is_flag=True, help='show debug output')
@click.option('--verbose', '-v', default=1, type=int, help='verbosity level 0-3')
@click.option('--interval', type=int, help='how may days does report cover')
@click.option('--date', default='', type=str, help='date in yyyy-mm-dd format')
@click.option('--title', default='', type=str, help='Title of report for emailing')
@click.option('--dump', is_flag=True, help='dump to screen do not email')
def main(debug, verbose, interval, date, title, dump):
    global dbg

    # load environmental variables
    load_dotenv(resource_path(".env"))

    if os.getenv('HELP'):
      with click.get_current_context() as ctx:
        click.echo(ctx.get_help())
        ctx.exit()

    debug = resolve_flag('DEBUG', debug)
    verbosity = resolve_int('VERBOSE', verbose)
    interval = resolve_int('INTERVAL', interval)
    date = resolve_text('DATE', date)
    title = resolve_text('DATE', title)
    dump = resolve_flag('DUMP', dump)

    if debug:
        dbg = verbose

    if date:
        date = datetime.strptime(date, '%Y-%m-%d')
    else:
        date = datetime.now()

    if not title:
        title = os.getenv('INTERVAL_TITLE')

    xlsfile = resource_path(os.getenv('XLSFILE'))

    report_date =  date
    report_start = date - timedelta(days=int(interval))

    try:
        oprs = fetch_oprs(report_start, report_date)
        if dump:
            dump_oprs(oprs)
            print()
        else:
            filename, longfilename = write_sheet(oprs, xlsfile, report_date)
            mail_results(
                filename[:-5],
                '<p>Here is the ' + title + ' OS OPR Sales Report.</p>',
                attachment = longfilename
            )
            os.remove(longfilename)


    except Exception as e:
        mail_results(
            'OS OPR Sales Processing Error',
            '<p>Spreadsheet can not be updated due to script error:<br />\n' + str(e) + '</p>'
        )

    sys.exit(0)

if __name__ == "__main__":
    main()
